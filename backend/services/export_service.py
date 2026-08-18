from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

from services import storage_service

HEADER_FILL = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 50)


def _write_table(ws, headers: list[str], rows: list[list], start_row: int = 1) -> int:
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    return start_row + len(rows) + 2


def build_xlsx(file_id: str) -> bytes:
    """Builds a multi-sheet Excel workbook: Overview, Column Types,
    Categorical Breakdown, Numeric Stats, and a Data Sample — for tabular
    files. For PDF-kind files, exports extracted headings and tables.
    """
    meta = storage_service.get_meta(file_id)
    wb = Workbook()

    if meta["kind"] == "tabular":
        from services import excel_service

        df = excel_service.read_tabular(meta["stored_path"])
        stats = storage_service.cache_get(file_id, "stats") or excel_service.analyze(df)

        ws = wb.active
        ws.title = "Overview"
        _write_table(
            ws,
            ["Metric", "Value"],
            [
                ["File name", meta["filename"]],
                ["Rows", stats["shape"]["rows"]],
                ["Columns", stats["shape"]["columns"]],
                ["Missing values", stats["missing_values"]["total"]],
                ["Missing %", stats["missing_values"]["percent_of_cells"]],
                ["Duplicate rows", stats["duplicate_rows"]],
            ],
        )
        _autosize(ws)

        ws2 = wb.create_sheet("Column Types")
        _write_table(ws2, ["Column", "Type"], list(stats["column_types"].items()))
        _autosize(ws2)

        if stats["categorical_stats"]:
            ws3 = wb.create_sheet("Categorical Breakdown")
            row = 1
            for col, items in stats["categorical_stats"].items():
                ws3.cell(row=row, column=1, value=col).font = Font(bold=True)
                row += 1
                row = _write_table(
                    ws3, ["Value", "Count"], [[it["value"], it["count"]] for it in items], row
                )
            _autosize(ws3)

        if stats["numeric_stats"]:
            ws4 = wb.create_sheet("Numeric Stats")
            headers = ["Column", "Count", "Mean", "Std", "Min", "Max", "Median", "Sum"]
            rows = [
                [col, s["count"], s["mean"], s["std"], s["min"], s["max"], s["median"], s["sum"]]
                for col, s in stats["numeric_stats"].items()
            ]
            _write_table(ws4, headers, rows)
            _autosize(ws4)

        ws5 = wb.create_sheet("Data Sample")
        sample = excel_service.sample_rows(df, n=100)
        if sample:
            _write_table(ws5, list(sample[0].keys()), [list(r.values()) for r in sample])
            _autosize(ws5)

    else:
        from services import pdf_service

        extraction = storage_service.cache_get(file_id, "extraction") or pdf_service.extract(
            meta["stored_path"]
        )
        ws = wb.active
        ws.title = "Headings"
        _write_table(
            ws, ["Page", "Heading"], [[h["page"], h["text"]] for h in extraction["headings"]]
        )
        _autosize(ws)

        if extraction["tables"]:
            for i, table in enumerate(extraction["tables"][:20], start=1):
                sheet = wb.create_sheet(f"Table {i} (p{table['page']})"[:31])
                for r, row in enumerate(table["rows"], start=1):
                    for c, val in enumerate(row, start=1):
                        sheet.cell(row=r, column=c, value=val)
                _autosize(sheet)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(file_id: str) -> bytes:
    """Builds a styled PDF report: file info, key stats table, and any
    cached AI insights."""
    meta = storage_service.get_meta(file_id)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "BasiraTitle", parent=styles["Title"], textColor=colors.HexColor("#111827")
    )
    heading_style = ParagraphStyle(
        "BasiraHeading", parent=styles["Heading2"], textColor=colors.HexColor("#16A34A")
    )
    body_style = styles["BodyText"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm, leftMargin=2 * cm, rightMargin=2 * cm
    )
    story = [
        Paragraph("Basira AI — File Report", title_style),
        Spacer(1, 4),
        Paragraph(meta["filename"], body_style),
        Spacer(1, 16),
    ]

    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]
    )

    if meta["kind"] == "tabular":
        from services import excel_service

        df = excel_service.read_tabular(meta["stored_path"])
        stats = storage_service.cache_get(file_id, "stats") or excel_service.analyze(df)

        story.append(Paragraph("Overview", heading_style))
        rows = [
            ["Metric", "Value"],
            ["Rows", str(stats["shape"]["rows"])],
            ["Columns", str(stats["shape"]["columns"])],
            ["Missing values", f"{stats['missing_values']['total']} ({stats['missing_values']['percent_of_cells']}%)"],
            ["Duplicate rows", str(stats["duplicate_rows"])],
        ]
        story.append(Table(rows, style=table_style, colWidths=[6 * cm, 6 * cm]))
        story.append(Spacer(1, 16))

        story.append(Paragraph("Column Types", heading_style))
        col_rows = [["Column", "Type"]] + [[k, v] for k, v in stats["column_types"].items()]
        story.append(Table(col_rows, style=table_style, colWidths=[8 * cm, 4 * cm]))
        story.append(Spacer(1, 16))
    else:
        from services import pdf_service

        extraction = storage_service.cache_get(file_id, "extraction") or pdf_service.extract(
            meta["stored_path"]
        )
        story.append(Paragraph("Document Overview", heading_style))
        story.append(Paragraph(f"Pages: {extraction['page_count']}", body_style))
        story.append(Spacer(1, 10))
        if extraction["headings"]:
            story.append(Paragraph("Outline", heading_style))
            for h in extraction["headings"][:30]:
                story.append(Paragraph(f"• {h['text']} (p.{h['page']})", body_style))
            story.append(Spacer(1, 16))

    insights = storage_service.cache_get(file_id, "insights")
    if insights:
        story.append(Paragraph("AI Insights", heading_style))
        if insights.get("summary"):
            story.append(Paragraph(insights["summary"], body_style))
            story.append(Spacer(1, 8))
        for point in insights.get("key_points", []):
            story.append(Paragraph(f"• {point}", body_style))

    doc.build(story)
    return buf.getvalue()
